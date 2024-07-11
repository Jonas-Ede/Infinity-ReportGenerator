from enum import Enum
import sys
import os
import openpyxl
import datetime
import CSVhandle
from openpyxl import *
from openpyxl.styles import PatternFill
from tkinter import * 
from tkinter import filedialog

def cell_to_datetime(cell_data):
    split = cell_data.split(" ")
    date = split[0].split("/")
    time = split[1].split(":")
    return datetime.datetime(int(date[2]), int(date[0]), int(date[1]), int(time[0]), int(time[1]))

def cell_to_time(cell_data):
    split = cell_data.split(" ")
    time = split[1].split(":")
    return time[0] + ":" + time[1]

days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

#Select Clockshark csv file
root = Tk()
root.title('Timesheet Report Generator')
root.filename = filedialog.askopenfilename(title="Select Clockshark CSV", filetypes=(("CSV Files", "*.csv"),))     
fileType = (root.filename).lower().split('.')[-1]

#ensure filetype is csv
if fileType not in ("csv"):
    print("Error: Incorrect File Type")
    sys.exit()
      
dataDFs = CSVhandle.csv_to_datArr(str(root.filename))

directory_path = r".\TEMPLATES"
folder = os.fsencode(directory_path)

for file in os.listdir(folder):
    filename = os.fsdecode(file)
    filepath = os.path.join(directory_path, filename)
    
    title = filename.split('_')[0]
    isCostSheet = (filename.split('_')[1]).upper() == "COST"
    
    
    if title in dataDFs:
        target_dataframe = dataDFs[title]
        print(f"Data for {title}:")
        for date in dataDFs[title]:
            df = dataDFs[title][date] 
            workbook_obj = openpyxl.load_workbook(filepath)
            report = workbook_obj.active
            print(df.shape)
            if(isCostSheet):
                # fill in cell 1M with the date and 3M with the day of week
                report['M1'] = str(date)
                report['M3'] = date.strftime('%A') 
                # Look through the second column of report to find a match for the FullName section of df, when match is found, 
                # fill in the 12th 13th 14th and 16th with in, out, time, and overtime hours from the Start and End Columns of df
                for index, row in df.iterrows():
                    hourval = row['Regular (Mins)']/60
                    if(hourval > 8):
                        Rtime = 8
                        Otime = (hourval-8) + (row['OverTime (Mins)']/60)
                    else:
                        Rtime = hourval
                        Otime = 0
                    for report_row in report.iter_rows(min_row=2, max_row=report.max_row, min_col=2, max_col=2):
                        cell_value = report_row[0].value
                        if cell_value == row['FullName'] or cell_value == row['FullNameNC']:
                            # Fill in cells 12, 13, 14, and 16 with data from df
                            report.cell(row=report_row[0].row, column=12, value=row['Start'].strftime('%H:%M'))
                            report.cell(row=report_row[0].row, column=13, value=cell_to_time(row['End']))
                            report.cell(row=report_row[0].row, column=14, value=Rtime)
                            report.cell(row=report_row[0].row, column=16, value=Otime)
                            break  # Exit the loop after finding the match
                workbook_obj.save(f"{title}-{str(date)}-Cost-Report.xlsx")
            else:
                report['E7'] = str(date)
                report['Q7'] = date.strftime('%A') 
                for index, row in df.iterrows():
                    print(row['FullNameNC'])
                    hourval = row['Regular (Mins)']/60
                    if(hourval > 8):
                        Rtime = 8
                        Otime = (hourval-8) + (row['OverTime (Mins)']/60)
                    else:
                        Rtime = hourval
                        Otime = 0
                    for report_row in report.iter_rows(min_row=3, max_row=report.max_row, min_col=3, max_col=3):
                        cell_value = report_row[0].value
                        if cell_value == row['FullName'] or cell_value == row['FullNameNC']:
                            report.cell(row=report_row[0].row, column=6, value=row['Start'].strftime('%H:%M'))
                            report.cell(row=report_row[0].row, column=7, value=cell_to_time(row['End']))
                            report.cell(row=report_row[0].row, column=8, value=Rtime)
                            report.cell(row=report_row[0].row, column=10, value=Otime)
                    for report_row in report.iter_rows(min_row=3, max_row=report.max_row, min_col=20, max_col=20):
                        cell_value = report_row[0].value
                        if cell_value == row['FullName'] or cell_value == row['FullNameNC']:
                            report.cell(row=report_row[0].row, column=23, value=row['Start'].strftime('%H:%M'))
                            report.cell(row=report_row[0].row, column=24, value=cell_to_time(row['End']))
                            report.cell(row=report_row[0].row, column=25, value=Rtime)
                            report.cell(row=report_row[0].row, column=27, value=Otime)
                workbook_obj.save(f"{title}-{str(date)}-Report.xlsx")   
    else:
        print(f"No data found for {title}.")
    
    
    

    
    
    
    
    
    
    
'''
dateYear = int(dateRaw[2])
dateMonth = int(dateRaw[0])
dateDay = int(dateRaw[1])
dateInstance = datetime.date(dateYear, dateMonth, dateDay)
dateDOW = days_of_week[dateInstance.weekday()]

report = newReport.active
costReport = newCostReport.active

#start with cost report
for person in datArr:
    timeIn = cell_to_datetime(person[3])
    timeOut = cell_to_datetime(person[4])
    time_difference = timeOut - timeIn
    clockedHours = time_difference.total_seconds() / 3600
    for r in range(22, 22+len(datArr)):
        if costReport.cell(row=r, column=2).value == person[0]:
            costReport.cell(row=r, column=12).value = timeIn.time()
            costReport.cell(row=r, column=13).value = timeOut.time()
            #weekends are all OT
            if(timeIn.weekday() >= 5):
                costReport.cell(row=r, column=14).value = 0
                costReport.cell(row=r, column=16).value = clockedHours
            #if OT and RT worked
            elif(clockedHours >= 8):
                costReport.cell(row=r, column=14).value = 8
                costReport.cell(row=r, column=16).value = clockedHours - 8
            #if only RT worked
            else:
                costReport.cell(row=r, column=14).value = clockedHours
                costReport.cell(row=r, column=16).value = 0
        
    
    for r in range(20, 22+len(datArr)):
        if report.cell(row=r, column=3).value == person[0]:
            report.cell(row=r, column=6).value = timeIn.time()
            report.cell(row=r, column=7).value = timeOut.time()
            #weekends are all OT
            if(timeIn.weekday() >= 5):
                report.cell(row=r, column=8).value = 0
                report.cell(row=r, column=10).value = clockedHours
            #if OT and RT worked
            elif(clockedHours >= 8):
                report.cell(row=r, column=8).value = 8
                report.cell(row=r, column=10).value = clockedHours - 8
            #if only RT worked
            else:
                report.cell(row=r, column=8).value = clockedHours
                report.cell(row=r, column=10).value = 0
        
        if report.cell(row=r, column=20).value == person[0]:
            report.cell(row=r, column=23).value = timeIn.time()
            report.cell(row=r, column=24).value = timeOut.time()
            #weekends are all OT
            if(timeIn.weekday() >= 5):
                report.cell(row=r, column=25).value = 0
                report.cell(row=r, column=27).value = clockedHours
            #if OT and RT worked
            elif(clockedHours >= 8):
                report.cell(row=r, column=25).value = 8
                report.cell(row=r, column=27).value = clockedHours - 8
            #if only RT worked
            else:
                report.cell(row=r, column=25).value = clockedHours
                report.cell(row=r, column=27).value = 0

for row in range(14,18):
    ogCell = (report.cell(row=row, column=1).value)
    ogCellsplit = str(ogCell).split(" ")
    name = ogCellsplit[2] + ", " + ogCellsplit[1]
    found = False
    for person in datArr:
        if person[0] == name:
            found = True
            report.cell(row=row, column=1).value = " ".join(ogCellsplit[0:5]) + " " + (cell_to_time(person[3])) + " - " + (cell_to_time(person[4]))



newReport.save(str(dateDOW) + "-" + str(dateYear) + "-" + str(dateMonth) + "-" + str(dateDay) + "-Report.xlsx")
newCostReport.save(str(dateDOW) + "-" + str(dateYear) + "-" + str(dateMonth) + "-" + str(dateDay) + "-CostReport.xlsx")
'''