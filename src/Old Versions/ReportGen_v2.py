from math import nan
import sys
import os
import openpyxl
import datetime
import pandas as pd
import CSVhandle
import pyodbc
from openpyxl import *
from tkinter import * 
from tkinter import filedialog
'''
task_table = {
    "Cleaning - Day Shift 1015-01": "COST CODE - 1 Day Shift Cleaning",
    "Cleaning - Night Shift 1015-02": "COST CODE - 2 Night Shift Cleaning",
    "Carpentry - Day Shift 1021-01": "COST CODE - 1 Day Shift Carpentry",
    "Carpentry - Night Shift 1021-11": "COST CODE - 11 Night Shift Carpentry",
    "Flying Equipment - Day Shift 1021-02": "COST CODE - 2 Day Shift Flying Equipment",
    "Flying Equipment - Night Shift 1021-12": "COST CODE - 12 Night Shift Flying Equipment",
    "De-Watering - Day Shift 1021-03": "COST CODE - 3 Day Shift De-Watering",
    "De-Watering - Night Shift 1021-13": "COST CODE - 13 Night Shift De-Watering",
    "Exterior Work - Day Shift 1021-04": "COST CODE - 4 Day Shift Exterior Work",
    "Exterior Work - Night Shift 1021-14": "COST CODE - 14 Night Shift Exterior Work",
    "Special Tasks - Day Shift 1021-05": "COST CODE - 5 Day Shift Special Tasks",
    "Special Tasks - Night Shift 1021-15": "COST CODE - 15 Night Shift Special Tasks"
}
'''
task_table = {
    "1015-01": "COST CODE - 1 Day Shift Cleaning",
    "1015-02": "COST CODE - 2 Night Shift Cleaning",
    "1021-01": "COST CODE - 1 Day Shift Carpentry",
    "1021-11": "COST CODE - 11 Night Shift Carpentry",
    "1021-02": "COST CODE - 2 Day Shift Flying Equipment",
    "1021-12": "COST CODE - 12 Night Shift Flying Equipment",
    "1021-03": "COST CODE - 3 Day Shift De-Watering",
    "1021-13": "COST CODE - 13 Night Shift De-Watering",
    "1021-04": "COST CODE - 4 Day Shift Exterior Work",
    "1021-14": "COST CODE - 14 Night Shift Exterior Work",
    "1021-05": "COST CODE - 5 Day Shift Special Tasks",
    "1021-15": "COST CODE - 15 Night Shift Special Tasks",
    "10087" : "COST CODE - 1 Day Shift Cleaning"
}

task_columns = {
    "COST CODE - 1 Day Shift Cleaning" : 6,
    "COST CODE - 2 Night Shift Cleaning" : 9,
    "COST CODE - 1 Day Shift Carpentry" : 12,
    "COST CODE - 11 Night Shift Carpentry" : 15,
    "COST CODE - 2 Day Shift Flying Equipment" : 18,
    "COST CODE - 12 Night Shift Flying Equipment" : 21,
    "COST CODE - 3 Day Shift De-Watering" : 24,
    "COST CODE - 13 Night Shift De-Watering" : 27,
    "COST CODE - 4 Day Shift Exterior Work" : 30,
    "COST CODE - 14 Night Shift Exterior Work" : 33,
    "COST CODE - 5 Day Shift Special Tasks" : 36,
    "COST CODE - 15 Night Shift Special Tasks" : 39
}


def cell_to_datetime(cell_data):
    split = cell_data.split(" ")
    date = split[0].split("/")
    time = split[1].split(":")
    return datetime.datetime(int(date[2]), int(date[0]), int(date[1]), int(time[0]), int(time[1]))

def cell_to_time(cell_data):
    split = cell_data.split(" ")
    time = split[1].split(":")
    return time[0] + ":" + time[1]

days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
SERVER = 'infinitysrv.database.windows.net'
DATABASE = 'infinitydb'
USERNAME = 'infinityadmin'
PASSWORD = 'Cranes24Canyon&'
connectionString = f'DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={SERVER};DATABASE={DATABASE};UID={USERNAME};PWD={PASSWORD}'
conn = pyodbc.connect(connectionString)

#Select Clockshark csv file
root = Tk()
root.title('Timesheet Report Generator')
root.filename = filedialog.askopenfilename(title="Select Clockshark CSV", filetypes=(("CSV Files", "*.csv"),))     
fileType = (root.filename).lower().split('.')[-1]

#ensure filetype is csv
if fileType not in ("csv"):
    print("Error: Incorrect File Type")
    sys.exit()
      
dataDFs = CSVhandle.csv_to_datArr(str(root.filename))

directory_path = r".\TEMPLATES"
folder = os.fsencode(directory_path)
breakdown_filename = os.fsdecode(r".\GENERAL_WORK_BREAKDOWN_TEMPLATE.xlsx")
filepath_breakdown = os.path.join(directory_path, breakdown_filename)
daily_filename = os.fsdecode(f".\\DAILY_REPORT_TEMPLATE_EMPTY.xlsx")
filepath_daily = os.path.join(directory_path, daily_filename)

for jobNum, day_dataframes in dataDFs.items():
    for date, df in day_dataframes.items():
        workbook_breakdown = openpyxl.load_workbook(filepath_breakdown)
        breakdown_report = workbook_breakdown.active
        if jobNum == "YC1015":
            breakdown_report['D1'] = str(date)
            breakdown_report['D3'] = date.strftime('%A')
            for report_row in breakdown_report.iter_rows(min_row=7, min_col=0, max_col=52, max_row=88, values_only=False):
                name = report_row[0].value
                df_row = df.query('FullName == @name')
                if not df_row.empty and task_table.get(df_row['TaskCode'].values[0]) != None:
                    timetot = 0
                    for i in range(len(df_row['TaskCode'].values)):
                        jobtime = df_row['Regular (Mins)'].values[i] / 60  # Convert minutes to hours
                        remaining_regular_time = max(0, 8 - timetot)
                        Rtime = min(jobtime, remaining_regular_time)
                        Otime = max(0, jobtime - Rtime)
                        timetot += jobtime
                        if task_columns.get(task_table.get(df_row['TaskCode'].values[i])) is not None:
                            if Rtime != 0: report_row[task_columns.get(task_table.get(df_row['TaskCode'].values[i])) + 1].value = Rtime
                            if Otime != 0: report_row[task_columns.get(task_table.get(df_row['TaskCode'].values[i])) + 2].value = Otime
                                                              
            workbook_breakdown.save(f"{jobNum}_{str(date)}_GENERAL_ACTIVITY_BREAKDOWN.xlsx") 
        
        if os.path.isfile:
            workbook_daily = openpyxl.load_workbook(filepath_daily)
            daily_report = workbook_daily.active
            daily_report['E7'] = str(date)
            daily_report['Q7'] = date.strftime('%A')
            daily_report['Y7'] = jobNum
            #FILLING IN DATA FOR MANAGMENT
            for report_row in daily_report.iter_rows(min_row=15, min_col=1, max_col=30, max_row=20, values_only=False):
                init_val = report_row[0].value
                if init_val != None and init_val != 'Title':
                    person = init_val.split(' - ')[0].split('. ', 1)[1]
                    CSName = CSVhandle.fetch_name_from_alias(conn,person)
                    if CSName == None:
                        df_row = df.query('FullNameNC == @person')
                    else:
                        df_row = df.query('FullNameNC == @CSName')
                    if not df_row.empty:
                        hourval = sum(df_row['Regular (Mins)'].values)/60
                        
                        if(hourval > 8):
                            Rtime = 8
                            Otime = (hourval-8)
                        else:
                            Rtime = hourval
                            Otime = 0
                        report_row[0].value = init_val + ' - ' + pd.Timestamp(df_row['Start'].values[0]).strftime('%H:%M') + '-' + pd.Timestamp(df_row['End'].values[0]).strftime('%H:%M')
                        report_row[24].value = Rtime
                        report_row[26].value = Otime     
            
            for report_row in daily_report.iter_rows(min_row=22, min_col=3, max_col=30, max_row=72, values_only=False):
                person1 = report_row[0].value
                person2 = report_row[17].value
                #LEFT COLUMN
                if person1 != None and person1 != 'Name':
                    CSName = CSVhandle.fetch_name_from_alias(conn,person1)
                    if CSName == None:
                        df_row = df.query('FullNameNC == @person1')
                    else:
                        df_row = df.query('FullNameNC == @CSName')
                    if not df_row.empty:
                        hourval = sum(df_row['Regular (Mins)'].values)/60
                        
                        if(hourval > 8):
                            Rtime = 8
                            Otime = hourval-8
                        else:
                            Rtime = hourval
                            Otime = 0       
                        report_row[3].value = pd.Timestamp(df_row['Start'].values[0]).strftime('%H:%M')
                        report_row[4].value = pd.Timestamp(df_row['End'].values[len(df_row['End'].values)]).strftime('%H:%M')
                        report_row[5].value = Rtime
                        report_row[7].value = Otime
                #RIGHT COLUMN
                if person2 != None and person2 != 'Name':
                    CSName = CSVhandle.fetch_name_from_alias(conn,person2)
                    if CSName == None:
                        df_row = df.query('FullNameNC == @person2')
                    else:
                        df_row = df.query('FullNameNC == @CSName')
                    if  not df_row.empty:
                        hourval = sum(df_row['Regular (Mins)'].values)/60
                        if(hourval > 8):
                            Rtime = 8
                            Otime = hourval-8 
                        else:
                            Rtime = hourval
                            Otime = 0
                           
                        report_row[20].value = pd.Timestamp(df_row['Start'].values[0]).strftime('%H:%M')
                        report_row[21].value = pd.Timestamp(df_row['End'].values[0]).strftime('%H:%M')
                        report_row[22].value = Rtime
                        report_row[24].value = Otime   
            workbook_daily.save(f"{jobNum}_{str(date)}_DAILY_REPORT.xlsx")   