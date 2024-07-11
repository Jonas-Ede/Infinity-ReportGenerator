from math import nan
import sys
import os
import openpyxl
import datetime
import pandas as pd
import CSVhandle
import pyodbc
import copy
from openpyxl import *
from tkinter import * 
from tkinter import filedialog

'''
Notes: Hours are missing becasue time needs to actually be by the time catergories not by hours, if someone works job a then job b then back to job a job a must pay the overtime
'''

task_table = {
    "1015-01": "COST CODE - 1 Day Shift Cleaning",
    "1015-02": "COST CODE - 2 Night Shift Cleaning",
    "1021-01": "COST CODE - 1 Day Shift Carpentry",
    "1021-11": "COST CODE - 11 Night Shift Carpentry",
    "1021-02": "COST CODE - 2 Day Shift Flying Equipment",
    "1021-12": "COST CODE - 12 Night Shift Flying Equipment",
    "1021-03": "COST CODE - 3 Day Shift De-Watering",
    "1021-13": "COST CODE - 13 Night Shift De-Watering",
    "1021-04": "COST CODE - 4 Day Shift Exterior Work",
    "1021-14": "COST CODE - 14 Night Shift Exterior Work",
    "1021-05": "COST CODE - 5 Day Shift Special Tasks",
    "1021-15": "COST CODE - 15 Night Shift Special Tasks",
    "10087" : "COST CODE - 1 Day Shift Cleaning"
}

task_columns = {
    "COST CODE - 1 Day Shift Cleaning" : 7,
    "COST CODE - 2 Night Shift Cleaning" : 10,
    "COST CODE - 1 Day Shift Carpentry" : 13,
    "COST CODE - 11 Night Shift Carpentry" : 16,
    "COST CODE - 2 Day Shift Flying Equipment" : 19,
    "COST CODE - 12 Night Shift Flying Equipment" : 22,
    "COST CODE - 3 Day Shift De-Watering" : 25,
    "COST CODE - 13 Night Shift De-Watering" : 28,
    "COST CODE - 4 Day Shift Exterior Work" : 31,
    "COST CODE - 14 Night Shift Exterior Work" : 34,
    "COST CODE - 5 Day Shift Special Tasks" : 37,
    "COST CODE - 15 Night Shift Special Tasks" : 40
}


def cell_to_datetime(cell_data):
    split = cell_data.split(" ")
    date = split[0].split("/")
    time = split[1].split(":")
    return datetime.datetime(int(date[2]), int(date[0]), int(date[1]), int(time[0]), int(time[1]))

def cell_to_time(cell_data):
    split = cell_data.split(" ")
    time = split[1].split(":")
    return time[0] + ":" + time[1]

def excel_letter(n):
    result = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

#SQL Server informaton
SERVER = 'infinitysrv.database.windows.net'
DATABASE = 'infinitydb'
USERNAME = 'infinityadmin'
PASSWORD = 'Cranes24Canyon&'
connectionString = f'DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={SERVER};DATABASE={DATABASE};UID={USERNAME};PWD={PASSWORD}'
conn = pyodbc.connect(connectionString)

#Select Clockshark csv file
root = Tk()
root.title('Timesheet Report Generator')
root.filename = filedialog.askopenfilename(title="Select Clockshark CSV", filetypes=(("CSV Files", "*.csv"),))     
fileType = (root.filename).lower().split('.')[-1]

#ensure filetype is csv
if fileType not in ("csv"):
    print("Error: Incorrect File Type")
    sys.exit()
      
dataDFs = CSVhandle.csv_to_datArr(str(root.filename))

directory_path = r".\TEMPLATES"
folder = os.fsencode(directory_path)
breakdown_filename = os.fsdecode(r".\GENERAL_WORK_BREAKDOWN_TEMPLATE_CLEAR.xlsx")
filepath_breakdown = os.path.join(directory_path, breakdown_filename)
daily_filename = os.fsdecode(f".\\DAILY_REPORT_TEMPLATE_CLEAR.xlsx")
filepath_daily = os.path.join(directory_path, daily_filename)

for jobNum, day_dataframes in dataDFs.items():
    for date, df in day_dataframes.items():
        #WORK BREAKDOWN REPORT
        if jobNum == "YC1015":
            workbook_breakdown = openpyxl.load_workbook(filepath_breakdown)
            breakdown_report = workbook_breakdown.active
            breakdown_report['D1'] = str(date)
            breakdown_report['D3'] = date.strftime('%A')
            used_names = set()
            iterator = 7
            for index, row in df.iterrows():
                CSName = row['FullNameNC']
                if CSName not in used_names:
                    used_names.add(CSName)
                    df_row = df.query('FullNameNC == @CSName')
                    db_row = CSVhandle.fetch_alias(conn, CSName)
                    
                    if db_row is None:
                        print(f"No DB row returned for [{CSName}]")
                    elif task_table.get(df_row['TaskCode'].values[0]) != None:
                        print(f'Adding {CSName} to Breakdown report')
                        print(db_row)
                        timetot = 0 
                        breakdown_report[f'A{iterator}'].value = db_row[0]
                        for i in range(len(df_row['TaskCode'].values)):
                            jobtime = df_row['Regular (Mins)'].values[i] / 60  # Convert minutes to hours
                            remaining_regular_time = max(0, 8 - timetot)
                            Rtime = min(jobtime, remaining_regular_time)
                            Otime = max(0, jobtime - Rtime)
                            timetot += jobtime
                            if task_columns.get(task_table.get(df_row['TaskCode'].values[i])) is not None:
                                if Rtime != 0: breakdown_report[f'{excel_letter(task_columns.get(task_table.get(df_row['TaskCode'].values[i]))+1)}{iterator}'].value = Rtime
                                if Otime != 0: breakdown_report[f'{excel_letter(task_columns.get(task_table.get(df_row['TaskCode'].values[i]))+2)}{iterator}'].value = Otime
                        iterator += 1
            
            workbook_breakdown.save(f"{jobNum}_{str(date)}_GENERAL_ACTIVITY_BREAKDOWN.xlsx")
        else:
            #DAILY REPORT
            workbook_daily = openpyxl.load_workbook(filepath_daily)
            daily_report = workbook_daily.active
            daily_report['E7'] = str(date)
            daily_report['Q7'] = date.strftime('%A')
            daily_report['Y7'] = jobNum
            #FILLING IN DATA FOR MANAGMENT
            used_names = set()
            for report_row in daily_report.iter_rows(min_row=15, min_col=1, max_col=30, max_row=20, values_only=False):
                init_val = report_row[0].value
                if init_val != None and init_val != 'Title':
                    person = init_val.split(' - ')[0].split('. ', 1)[1]
                    result = CSVhandle.fetch_alias(conn,person)
                    CSName = result[0] if result != None else None
                    used_names.add(CSName if CSName != None else person)
                    if CSName == None:
                        df_row = df.query('FullNameNC == @person')
                    else:
                        df_row = df.query('FullNameNC == @CSName')
                    if not df_row.empty:
                        hourval = sum(df_row['Regular (Mins)'].values)/60
                        
                        if(hourval > 8):
                            Rtime = 8
                            Otime = (hourval-8)
                        else:
                            Rtime = hourval
                            Otime = 0
                        report_row[0].value = init_val + ' - ' + pd.Timestamp(df_row['Start'].values[0]).strftime('%H:%M') + '-' + pd.Timestamp(df_row['End'].values[0]).strftime('%H:%M')
                        report_row[24].value = Rtime
                        report_row[26].value = Otime     
                        
            row_trackers = {
                "AM Skilled Labor": 0,
                "AM Semi-Skilled Labor": 0,
                "PM Skilled Labor": 0,
                "PM Semi-Skilled Labor": 0,
                "PM Safety MGr": 0,
                "PM Supervisor": 0
            }

            isLeft_flags = {
                "AM Skilled Labor": True,
                "AM Semi-Skilled Labor": True,
                "PM Skilled Labor": True,
                "PM Semi-Skilled Labor": True,
                "PM Safety MGr": True,
                "PM Supervisor": True
            }

            # Find the starting row for each section
            for row_num, row in enumerate(daily_report.iter_rows(min_row=1, max_row=100, min_col=1, max_col=1), start=1):
                if row[0].value == "MANAGEMENT ON-SITE ":
                    mangiterator = row_num + 1
                elif row[0].value == "AM Skilled Labor":
                    row_trackers["AM Skilled Labor"] = row_num
                elif row[0].value == "AM Semi-Skilled  Labor":
                    row_trackers["AM Semi-Skilled Labor"] = row_num
                elif row[0].value == "PM Skilled Labor":
                    row_trackers["PM Skilled Labor"] = row_num
                elif row[0].value == "PM Semi-Skilled  Labor":
                    row_trackers["PM Semi-Skilled Labor"] = row_num
                elif row[0].value == "PM Safety MGr":
                    row_trackers["PM Safety MGr"] = row_num
                elif row[0].value == "PM Supervisor":
                    row_trackers["PM Supervisor"] = row_num
            
            for index, row in df.iterrows():
                CSName = row['FullNameNC']
                if CSName not in used_names:
                    used_names.add(CSName)
                    df_row = df.query('FullNameNC == @CSName')
                    db_row = CSVhandle.fetch_alias(conn, CSName)
                    if db_row is None:
                        print(f"No DB row returned for [{CSName}]")
                    else:
                        # Determine the labor type and get the corresponding row tracker
                        labor_type = db_row[4]
                        if labor_type == "AM SKILLED LABOR" or labor_type == "AM SAFETY MANAGER" or labor_type == "AM SUPERVISION":
                            row_tracker_key = "AM Skilled Labor"
                        elif labor_type == "AM SEMI SKILLED LABOR":
                            row_tracker_key = "AM Semi-Skilled Labor"
                        elif labor_type == "PM SKILLED LABOR":
                            row_tracker_key = "PM Skilled Labor"
                        elif labor_type == "PM SEMI SKILLED LABOR":
                            row_tracker_key = "PM Semi-Skilled Labor"
                        elif labor_type == "PM SAFETY MANAGER":
                            row_tracker_key = "PM Safety MGr"
                        elif labor_type == "PM SUPERVISION":
                            row_tracker_key = "PM Supervisor"
                        else:
                            continue  # Skip unknown labor types

                        iterator = row_trackers[row_tracker_key]
                        isLeft = isLeft_flags[row_tracker_key]

                        # Calculate regular and overtime hours
                        hourval = sum(df_row['Regular (Mins)'].values) / 60
                        if hourval > 8:
                            Rtime = 8
                            Otime = hourval - 8
                        else:
                            Rtime = hourval
                            Otime = 0

                        if isLeft:
                            isLeft_flags[row_tracker_key] = False
                            daily_report[f'C{iterator}'].value = db_row[1]
                            daily_report[f'F{iterator}'].value = pd.Timestamp(df_row['Start'].values[0]).strftime('%H:%M')
                            daily_report[f'G{iterator}'].value = pd.Timestamp(df_row['End'].values[-1]).strftime('%H:%M')
                            daily_report[f'H{iterator}'].value = Rtime
                            daily_report[f'J{iterator}'].value = Otime
                            daily_report[f'L{iterator}'].value = db_row[2] + " - " + ("AM" if db_row[3] == "Day" else "PM")
                        else:
                            isLeft_flags[row_tracker_key] = True
                            daily_report[f'T{iterator}'].value = db_row[1]
                            daily_report[f'W{iterator}'].value = pd.Timestamp(df_row['Start'].values[0]).strftime('%H:%M')
                            daily_report[f'X{iterator}'].value = pd.Timestamp(df_row['End'].values[-1]).strftime('%H:%M')
                            daily_report[f'Y{iterator}'].value = Rtime
                            daily_report[f'AA{iterator}'].value = Otime
                            daily_report[f'AC{iterator}'].value = db_row[2] + " - " + ("AM" if db_row[3] == "Day" else "PM")
                            row_trackers[row_tracker_key] += 1  # Increment row tracker for the labor type

            workbook_daily.save(f"{jobNum}_{str(date)}_DAILY_REPORT.xlsx")

