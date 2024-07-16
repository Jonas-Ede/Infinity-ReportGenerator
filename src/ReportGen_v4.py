from math import nan
import sys
import os
import openpyxl
import datetime
import pandas as pd
import CSVhandle
import pyodbc
import numpy as np
from openpyxl import *
from tkinter import * 
from tkinter import filedialog
from tqdm import tqdm
'''
Notes: Hours are missing becasue time needs to actually be by the time catergories not by hours, if someone works job a then job b then back to job a job a must pay the overtime
'''

task_table = {
    "1015-01": "COST CODE - 1 Day Shift Cleaning",
    "1015-02": "COST CODE - 2 Night Shift Cleaning",
    "1021-01": "COST CODE - 1 Day Shift Carpentry",
    "1021-11": "COST CODE - 11 Night Shift Carpentry",
    "1021-02": "COST CODE - 2 Day Shift Flying Equipment",
    "1021-12": "COST CODE - 12 Night Shift Flying Equipment",
    "1021-03": "COST CODE - 3 Day Shift De-Watering",
    "1021-13": "COST CODE - 13 Night Shift De-Watering",
    "1021-04": "COST CODE - 4 Day Shift Exterior Work",
    "1021-14": "COST CODE - 14 Night Shift Exterior Work",
    "1021-05": "COST CODE - 5 Day Shift Special Tasks",
    "1021-15": "COST CODE - 15 Night Shift Special Tasks",
    "10087" : "COST CODE - 1 Day Shift Cleaning"
}

task_columns = {
    "COST CODE - 1 Day Shift Cleaning" : 7,
    "COST CODE - 2 Night Shift Cleaning" : 10,
    "COST CODE - 1 Day Shift Carpentry" : 13,
    "COST CODE - 11 Night Shift Carpentry" : 16,
    "COST CODE - 2 Day Shift Flying Equipment" : 19,
    "COST CODE - 12 Night Shift Flying Equipment" : 22,
    "COST CODE - 3 Day Shift De-Watering" : 25,
    "COST CODE - 13 Night Shift De-Watering" : 28,
    "COST CODE - 4 Day Shift Exterior Work" : 31,
    "COST CODE - 14 Night Shift Exterior Work" : 34,
    "COST CODE - 5 Day Shift Special Tasks" : 37,
    "COST CODE - 15 Night Shift Special Tasks" : 40
}

def cell_to_datetime(cell_data):
    split = cell_data.split(" ")
    date = split[0].split("/")
    time = split[1].split(":")
    return datetime.datetime(int(date[2]), int(date[0]), int(date[1]), int(time[0]), int(time[1]))

def cell_to_time(cell_data):
    split = cell_data.split(" ")
    time = split[1].split(":")
    return time[0] + ":" + time[1]

def excel_letter(n):
    result = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

#SQL Server informaton
SERVER = 'infinitysrv.database.windows.net'
DATABASE = 'infinitydb'
USERNAME = 'infinityadmin'
PASSWORD = 'Cranes24Canyon&'
connectionString = f'DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={SERVER};DATABASE={DATABASE};UID={USERNAME};PWD={PASSWORD}'
conn = pyodbc.connect(connectionString)

#Select Clockshark csv file
root = Tk()
root.title('Timesheet Report Generator')
root.filename = filedialog.askopenfilename(title="Select Clockshark CSV", filetypes=(("CSV Files", "*.csv"),))     
fileType = (root.filename).lower().split('.')[-1]

#ensure filetype is csv
if fileType not in ("csv"):
    print("Error: Incorrect File Type")
    sys.exit()

#nothing = CSVhandle.csv_to_datArr(str(root.filename))      
dataDFs = CSVhandle.csv_to_datArr2(str(root.filename))

directory_path = r".\Infinity-ReportGenerator\TEMPLATES"
folder = os.fsencode(directory_path)
breakdown_filename = os.fsdecode(r".\GENERAL_WORK_BREAKDOWN_TEMPLATE_CLEAR.xlsx")
filepath_breakdown = os.path.join(directory_path, breakdown_filename)
daily_filename = os.fsdecode(f".\\DAILY_REPORT_TEMPLATE_CLEAR.xlsx")
filepath_daily = os.path.join(directory_path, daily_filename)

row_trackers = {
    "AM SKILLED LABOR": 0,
    "AM SEMI SKILLED LABOR ": 0,
    "PM SKILLED LABOR": 0,
    "PM SEMI SKILLED LABOR": 0,
    "PM SAFETY MANAGER": 0,
    "PM SUPERVISION": 0,
    "MANAGEMENT": 1
}

# Find the starting row for each section
template_wb = openpyxl.load_workbook(filepath_daily)
daily_report = template_wb.worksheets[1]
for row_num, row in enumerate(daily_report.iter_rows(min_row=1, max_row=100, min_col=1, max_col=1), start=1):
    if row[0].value == "MANAGEMENT ON-SITE ":
        mangiterator = row_num + 1
    elif row[0].value == "AM Skilled Labor":
        row_trackers["AM SKILLED LABOR"] = row_num
    elif row[0].value == "AM Semi-Skilled  Labor":
        row_trackers["AM SEMI SKILLED LABOR "] = row_num
    elif row[0].value == "PM Skilled Labor":
        row_trackers["PM SKILLED LABOR"] = row_num
    elif row[0].value == "PM Semi-Skilled  Labor":
        row_trackers["PM SEMI SKILLED LABOR"] = row_num
    elif row[0].value == "PM Safety MGr":
        row_trackers["PM SAFETY MANAGER"] = row_num
    elif row[0].value == "PM Supervisor":
        row_trackers["PM SUPERVISION"] = row_num

#print(row_trackers)

for date, df in dataDFs.items():
    CSVhandle.db_clear_times(conn)
    df = df.sort_values(by='Start')
    #This separates out all the different jobs and creates a different sheet for each
    unique_jobs = df["JobNumber"].unique().tolist()
    template_wb = openpyxl.load_workbook(filepath_daily)
    first_sheet = template_wb.worksheets[1]
    second_sheet = template_wb.worksheets[2]
    for job in unique_jobs:
        if job != 'YC1015':
            new_sheet = template_wb.copy_worksheet(first_sheet)
            new_sheet.title = job
            new_sheet['E7'] = str(date)
            new_sheet['Q7'] = date.strftime('%A')
            new_sheet['Y7'] = job
        else:
            new_sheet = template_wb.copy_worksheet(second_sheet)
            new_sheet.title = job
            new_sheet['D1'] = str(date)
            new_sheet['D3'] = date.strftime('%A')
            
    template_wb.remove(template_wb.worksheets[0])
    template_wb.remove(template_wb.worksheets[0])
    template_wb.remove(template_wb.worksheets[0])
        
    #for each row, pull their row from the server, use the stored overtime and regular time horus to fill in the respective sheet
    #DB_ROW HEADERS[CLockSharkName, SECAIAlias, Position, Shift, Labor Category, Craft, RegularTimeLog, OverTimeLog]
    for index, df_row in tqdm(df.iterrows(), desc="Processing People Data..."):
        CSName = df_row['FullNameNC']
        db_row = CSVhandle.fetch_alias(conn, CSName)
        
        if db_row == None:
            print(f'Database row for {CSName} does not exist')
        else:
            #calculate worked hours and update hours in database    
            hourval = df_row['Regular (Mins)']/60
            rt_worked = db_row[8]
            ot_worked = db_row[9]
            remaining_rt_hours = 8 - rt_worked
            if hourval <= remaining_rt_hours:
                rt_worked += hourval
                added_rt = hourval
                added_ot = 0
            else:
                rt_worked = 8
                ot_worked += (hourval - remaining_rt_hours)
                added_rt = remaining_rt_hours
                added_ot = (hourval - remaining_rt_hours) 
            CSVhandle.db_update_times(conn, db_row[0], rt_worked, ot_worked)
            
            active_sheet = template_wb.worksheets[unique_jobs.index(df_row['JobNumber'])]
            if df_row['JobNumber'] != 'YC1015':
                #print('Adding to daily report')
                if db_row[4] == 'AM MANAGEMENT':
                    iter = 15
                    while(active_sheet[f'A{iter}'].value != None):
                        iter+=1
                    active_sheet[f'A{iter}'].value = f'{row_trackers['MANAGEMENT']}. {db_row[1]} - {db_row[2]} - {pd.Timestamp(df_row['Start'].value).strftime('%H:%M')}-{pd.Timestamp(df_row['End'].value).strftime('%H:%M')}'
                    active_sheet[f'Y{iter}'].value = added_rt
                    active_sheet[f'AA{iter}'].value = added_ot
                    row_trackers['MANAGEMENT'] += 1

                else:
                    row_key = db_row[4] if db_row[4] != 'AM SUPERVISION' and db_row[4] != 'AM SAFETY MANAGER' else "AM SKILLED LABOR"
                    iter = row_trackers[row_key]
                    n1 = active_sheet[f'C{iter}'].value
                    n2 = active_sheet[f'T{iter}'].value
                    while(n1 != None and n2 != None):
                        if(n1 == db_row[1]):
                            active_sheet[f'G{iter}'].value = pd.Timestamp(df_row['End'].value).strftime('%H:%M')
                            active_sheet[f'H{iter}'].value = active_sheet[f'H{iter}'].value + added_rt
                            active_sheet[f'J{iter}'].value = active_sheet[f'J{iter}'].value + added_ot
                            break
                        elif(n2 == db_row[1]):
                            active_sheet[f'X{iter}'].value = pd.Timestamp(df_row['End'].value).strftime('%H:%M')
                            active_sheet[f'Y{iter}'].value = active_sheet[f'Y{iter}'].value + added_rt
                            active_sheet[f'AA{iter}'].value = active_sheet[f'AA{iter}'].value + added_ot
                            break
                        else:
                            iter+=1
                            n1 = active_sheet[f'C{iter}'].value
                            n2 = active_sheet[f'T{iter}'].value
                    if n1 == None:
                        active_sheet[f'C{iter}'].value = db_row[1]
                        active_sheet[f'F{iter}'].value = pd.Timestamp(df_row['Start'].value).strftime('%H:%M')
                        active_sheet[f'G{iter}'].value = pd.Timestamp(df_row['End'].value).strftime('%H:%M')
                        active_sheet[f'H{iter}'].value = added_rt
                        active_sheet[f'J{iter}'].value = added_ot
                        active_sheet[f'L{iter}'].value = db_row[2] + " - " + ("AM" if db_row[3] == "Day" else "PM")
                    elif n2 == None:
                        active_sheet[f'T{iter}'].value = db_row[1]
                        active_sheet[f'W{iter}'].value = pd.Timestamp(df_row['Start'].value).strftime('%H:%M')
                        active_sheet[f'X{iter}'].value = pd.Timestamp(df_row['End'].value).strftime('%H:%M')
                        active_sheet[f'Y{iter}'].value = added_rt
                        active_sheet[f'AA{iter}'].value = added_ot
                        active_sheet[f'AC{iter}'].value = db_row[2] + " - " + ("AM" if db_row[3] == "Day" else "PM")       
            else:
                if(task_table.get(df_row['TaskCode']) != None):
                    iter = 7
                    while(active_sheet[f'A{iter}'].value != None):
                        if(active_sheet[f'A{iter}'].value == df_row["FullName"]):
                            active_sheet[f'{excel_letter(task_columns.get(task_table.get(df_row['TaskCode']))+1)}{iter}'].value = added_rt if active_sheet[f'{excel_letter(task_columns.get(task_table.get(df_row['TaskCode']))+1)}{iter}'].value == None else added_rt + float(active_sheet[f'{excel_letter(task_columns.get(task_table.get(df_row['TaskCode']))+1)}{iter}'].value)
                            active_sheet[f'{excel_letter(task_columns.get(task_table.get(df_row['TaskCode']))+2)}{iter}'].value = added_ot if active_sheet[f'{excel_letter(task_columns.get(task_table.get(df_row['TaskCode']))+2)}{iter}'].value == None else added_ot + float(active_sheet[f'{excel_letter(task_columns.get(task_table.get(df_row['TaskCode']))+2)}{iter}'].value)
                            break
                        else:
                            iter +=1 
                    if(active_sheet[f'A{iter}'].value == None):
                        active_sheet[f'A{iter}'].value = df_row["FullName"]
                        active_sheet[f'B{iter}'].value = db_row[5]
                        active_sheet[f'D{iter}'].value = db_row[2]
                        active_sheet[f'E{iter}'].value = float(db_row[6])
                        active_sheet[f'F{iter}'].value = float(db_row[7])
                        active_sheet[f'{excel_letter(task_columns.get(task_table.get(df_row['TaskCode']))+1)}{iter}'].value = added_rt
                        active_sheet[f'{excel_letter(task_columns.get(task_table.get(df_row['TaskCode']))+2)}{iter}'].value = added_ot
                else:
                    print(f'{CSName} does not have an asssigned task code, adding their hours to Special Tasks - Day Shift')
                    iter = 7
                    while(active_sheet[f'A{iter}'].value != None):
                        if(active_sheet[f'A{iter}'].value == df_row["FullName"]):
                            active_sheet[f'{excel_letter(task_columns.get(task_table.get('1021-05'))+1)}{iter}'].value = added_rt if active_sheet[f'{excel_letter(task_columns.get(task_table.get('1021-05'))+1)}{iter}'].value == None else added_rt + float(active_sheet[f'{excel_letter(task_columns.get(task_table.get('1021-05'))+1)}{iter}'].value)
                            active_sheet[f'{excel_letter(task_columns.get(task_table.get('1021-05'))+2)}{iter}'].value = added_ot if active_sheet[f'{excel_letter(task_columns.get(task_table.get('1021-05'))+2)}{iter}'].value == None else added_ot + float(active_sheet[f'{excel_letter(task_columns.get(task_table.get('1021-05'))+2)}{iter}'].value)
                            break
                        else:
                            iter +=1 
                    if(active_sheet[f'A{iter}'].value == None):
                        active_sheet[f'A{iter}'].value = df_row["FullName"]
                        active_sheet[f'B{iter}'].value = db_row[5]
                        active_sheet[f'D{iter}'].value = db_row[2]
                        active_sheet[f'E{iter}'].value = float(db_row[6])
                        active_sheet[f'F{iter}'].value = float(db_row[7])
                        active_sheet[f'{excel_letter(task_columns.get(task_table.get('1021-05'))+1)}{iter}'].value = added_rt
                        active_sheet[f'{excel_letter(task_columns.get(task_table.get('1021-05'))+2)}{iter}'].value = added_ot

    CSVhandle.db_clear_times(conn)
    template_wb.save(f"{str(date)}_REPORTS.xlsx")

'''
for jobNum, day_dataframes in dataDFs.items():
    for date, df in day_dataframes.items():
        #WORK BREAKDOWN REPORT
        if jobNum == "YC1015":
            workbook_breakdown = openpyxl.load_workbook(filepath_breakdown)
            breakdown_report = workbook_breakdown.active
            breakdown_report['D1'] = str(date)
            breakdown_report['D3'] = date.strftime('%A')
            used_names = set()
            iterator = 7
            for index, row in df.iterrows():
                CSName = row['FullNameNC']
                if CSName not in used_names:
                    used_names.add(CSName)
                    df_row = df.query('FullNameNC == @CSName')
                    db_row = CSVhandle.fetch_alias(conn, CSName)
                    
                    if db_row is None:
                        print(f"No DB row returned for [{CSName}]")
                    elif task_table.get(df_row['TaskCode'].values[0]) != None:
                        print(f'Adding {CSName} to Breakdown report')
                        print(db_row)
                        timetot = 0 
                        breakdown_report[f'A{iterator}'].value = db_row[0]
                        for i in range(len(df_row['TaskCode'].values)):
                            jobtime = df_row['Regular (Mins)'].values[i] / 60  # Convert minutes to hours
                            remaining_regular_time = max(0, 8 - timetot)
                            Rtime = min(jobtime, remaining_regular_time)
                            Otime = max(0, jobtime - Rtime)
                            timetot += jobtime
                            if task_columns.get(task_table.get(df_row['TaskCode'].values[i])) is not None:
                                if Rtime != 0: breakdown_report[f'{excel_letter(task_columns.get(task_table.get(df_row['TaskCode'].values[i]))+1)}{iterator}'].value = Rtime
                                if Otime != 0: breakdown_report[f'{excel_letter(task_columns.get(task_table.get(df_row['TaskCode'].values[i]))+2)}{iterator}'].value = Otime
                        iterator += 1
            
            workbook_breakdown.save(f"{jobNum}_{str(date)}_GENERAL_ACTIVITY_BREAKDOWN.xlsx")
        else:
            #DAILY REPORT
            workbook_daily = openpyxl.load_workbook(filepath_daily)
            daily_report = workbook_daily.active
            daily_report['E7'] = str(date)
            daily_report['Q7'] = date.strftime('%A')
            daily_report['Y7'] = jobNum
            #FILLING IN DATA FOR MANAGMENT
            used_names = set()
            for report_row in daily_report.iter_rows(min_row=15, min_col=1, max_col=30, max_row=20, values_only=False):
                init_val = report_row[0].value
                if init_val != None and init_val != 'Title':
                    person = init_val.split(' - ')[0].split('. ', 1)[1]
                    result = CSVhandle.fetch_alias(conn,person)
                    CSName = result[0] if result != None else None
                    used_names.add(CSName if CSName != None else person)
                    if CSName == None:
                        df_row = df.query('FullNameNC == @person')
                    else:
                        df_row = df.query('FullNameNC == @CSName')
                    if not df_row.empty:
                        hourval = sum(df_row['Regular (Mins)'].values)/60
                        
                        if(hourval > 8):
                            Rtime = 8
                            Otime = (hourval-8)
                        else:
                            Rtime = hourval
                            Otime = 0
                        report_row[0].value = init_val + ' - ' + pd.Timestamp(df_row['Start'].values[0]).strftime('%H:%M') + '-' + pd.Timestamp(df_row['End'].values[0]).strftime('%H:%M')
                        report_row[24].value = Rtime
                        report_row[26].value = Otime     
                        
            row_trackers = {
                "AM Skilled Labor": 0,
                "AM Semi-Skilled Labor": 0,
                "PM Skilled Labor": 0,
                "PM Semi-Skilled Labor": 0,
                "PM Safety MGr": 0,
                "PM Supervisor": 0
            }

            isLeft_flags = {
                "AM Skilled Labor": True,
                "AM Semi-Skilled Labor": True,
                "PM Skilled Labor": True,
                "PM Semi-Skilled Labor": True,
                "PM Safety MGr": True,
                "PM Supervisor": True
            }

            # Find the starting row for each section
            for row_num, row in enumerate(daily_report.iter_rows(min_row=1, max_row=100, min_col=1, max_col=1), start=1):
                if row[0].value == "MANAGEMENT ON-SITE ":
                    mangiterator = row_num + 1
                elif row[0].value == "AM Skilled Labor":
                    row_trackers["AM Skilled Labor"] = row_num
                elif row[0].value == "AM Semi-Skilled  Labor":
                    row_trackers["AM Semi-Skilled Labor"] = row_num
                elif row[0].value == "PM Skilled Labor":
                    row_trackers["PM Skilled Labor"] = row_num
                elif row[0].value == "PM Semi-Skilled  Labor":
                    row_trackers["PM Semi-Skilled Labor"] = row_num
                elif row[0].value == "PM Safety MGr":
                    row_trackers["PM Safety MGr"] = row_num
                elif row[0].value == "PM Supervisor":
                    row_trackers["PM Supervisor"] = row_num
            
            for index, row in df.iterrows():
                CSName = row['FullNameNC']
                if CSName not in used_names:
                    used_names.add(CSName)
                    df_row = df.query('FullNameNC == @CSName')
                    db_row = CSVhandle.fetch_alias(conn, CSName)
                    if db_row is None:
                        print(f"No DB row returned for [{CSName}]")
                    else:
                        # Determine the labor type and get the corresponding row tracker
                        labor_type = db_row[4]
                        if labor_type == "AM SKILLED LABOR" or labor_type == "AM SAFETY MANAGER" or labor_type == "AM SUPERVISION":
                            row_tracker_key = "AM Skilled Labor"
                        elif labor_type == "AM SEMI SKILLED LABOR":
                            row_tracker_key = "AM Semi-Skilled Labor"
                        elif labor_type == "PM SKILLED LABOR":
                            row_tracker_key = "PM Skilled Labor"
                        elif labor_type == "PM SEMI SKILLED LABOR":
                            row_tracker_key = "PM Semi-Skilled Labor"
                        elif labor_type == "PM SAFETY MANAGER":
                            row_tracker_key = "PM Safety MGr"
                        elif labor_type == "PM SUPERVISION":
                            row_tracker_key = "PM Supervisor"
                        else:
                            continue  # Skip unknown labor types

                        iterator = row_trackers[row_tracker_key]
                        isLeft = isLeft_flags[row_tracker_key]

                        # Calculate regular and overtime hours
                        hourval = sum(df_row['Regular (Mins)'].values) / 60
                        if hourval > 8:
                            Rtime = 8
                            Otime = hourval - 8
                        else:
                            Rtime = hourval
                            Otime = 0

                        if isLeft:
                            isLeft_flags[row_tracker_key] = False
                            daily_report[f'C{iterator}'].value = db_row[1]
                            daily_report[f'F{iterator}'].value = pd.Timestamp(df_row['Start'].values[0]).strftime('%H:%M')
                            daily_report[f'G{iterator}'].value = pd.Timestamp(df_row['End'].values[-1]).strftime('%H:%M')
                            daily_report[f'H{iterator}'].value = Rtime
                            daily_report[f'J{iterator}'].value = Otime
                            daily_report[f'L{iterator}'].value = db_row[2] + " - " + ("AM" if db_row[3] == "Day" else "PM")
                        else:
                            isLeft_flags[row_tracker_key] = True
                            daily_report[f'T{iterator}'].value = db_row[1]
                            daily_report[f'W{iterator}'].value = pd.Timestamp(df_row['Start'].values[0]).strftime('%H:%M')
                            daily_report[f'X{iterator}'].value = pd.Timestamp(df_row['End'].values[-1]).strftime('%H:%M')
                            daily_report[f'Y{iterator}'].value = Rtime
                            daily_report[f'AA{iterator}'].value = Otime
                            daily_report[f'AC{iterator}'].value = db_row[2] + " - " + ("AM" if db_row[3] == "Day" else "PM")
                            row_trackers[row_tracker_key] += 1  # Increment row tracker for the labor type

            workbook_daily.save(f"{jobNum}_{str(date)}_DAILY_REPORT.xlsx")

'''

